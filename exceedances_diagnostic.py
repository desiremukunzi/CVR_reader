"""
Diagnostic Tool for Exceedances Extraction
Run this to see exactly what's being extracted from your Excel files
"""

import openpyxl
from openpyxl import load_workbook
from pathlib import Path

def safe_get_cell_value(sheet, cell_ref, value_type='int'):
    """Safely get cell value with type conversion"""
    try:
        value = sheet[cell_ref].value
        
        if value is None or value == '' or value == '-':
            return 0 if value_type in ['int', 'float'] else None
        
        if value_type == 'int':
            return int(float(value))
        elif value_type == 'float':
            return float(value)
        elif value_type == 'str':
            return str(value).strip()
        else:
            return value
            
    except (ValueError, TypeError, AttributeError) as e:
        print(f"  ⚠️  Error getting value from {cell_ref}: {e}")
        return 0 if value_type in ['int', 'float'] else None

def diagnose_exceedances(file_path):
    """Diagnose exceedance extraction from an Excel file"""
    print("=" * 80)
    print(f"DIAGNOSING: {Path(file_path).name}")
    print("=" * 80)
    
    try:
        workbook = load_workbook(file_path, read_only=True, data_only=True)
        
        # Check if Summary sheet exists
        if 'Summary' not in workbook.sheetnames:
            print("❌ ERROR: 'Summary' sheet not found!")
            print(f"Available sheets: {workbook.sheetnames}")
            workbook.close()
            return
        
        summary = workbook['Summary']
        print("✅ Summary sheet found\n")
        
        # Check continuous exceedances (B9-C17)
        print("📊 CONTINUOUS EXCEEDANCES (B9-C17):")
        print("-" * 80)
        print(f"{'Cell':<8} {'Parameter (B col)':<20} {'Count (C col)':<15} {'Status'}")
        print("-" * 80)
        
        continuous_params = {
            'B9': 'IAS', 'B10': 'Alt', 'B11': 'Roll', 'B12': 'PITCH',
            'B13': 'Fcp', 'B14': 'N1/N2 Split', 'B15': 'N1', 'B16': 'N2', 'B17': 'Nmr',
        }
        
        continuous_found = 0
        for b_cell, expected_param in continuous_params.items():
            # Get parameter name from column B
            param_name = safe_get_cell_value(summary, b_cell, 'str')
            
            # Get count from column C (same row)
            row = b_cell[1:]  # Extract row number
            c_cell = f'C{row}'
            count = safe_get_cell_value(summary, c_cell, 'int')
            
            status = "✅" if count > 0 else "  "
            print(f"{b_cell:<8} {param_name or 'N/A':<20} {count:<15} {status}")
            
            if count > 0:
                continuous_found += 1
        
        print(f"\nContinuous exceedances found: {continuous_found}/9")
        
        # Check discrete exceedances (H3-I25)
        print("\n📊 DISCRETE EXCEEDANCES (H3-I25):")
        print("-" * 80)
        print(f"{'Cell':<8} {'Parameter (H col)':<20} {'Count (I col)':<15} {'Status'}")
        print("-" * 80)
        
        discrete_params = {
            'H3': 'iAPr/p', 'H4': 'iChips', 'H5': 'iEMG1', 'H6': 'iEMG2',
            'H7': 'iF_gen1', 'H8': 'iF_gen2', 'H9': 'iF_pump1', 'H10': 'iF_pump2',
            'H11': 'iF_pumpS', 'H12': 'iFire_KO-50', 'H13': 'iFire_mgb', 'H14': 'iFire_v1',
            'H15': 'iFire_v2', 'H16': 'iFire1', 'H17': 'iFire2', 'H18': 'iHSaux',
            'H19': 'iHSmain', 'H20': 'inFT1', 'H21': 'inFT2', 'H22': 'iOP_mgb',
            'H23': 'iOP1', 'H24': 'iOP2', 'H25': 'iQTmin',
        }
        
        discrete_found = 0
        for h_cell, expected_param in discrete_params.items():
            # Get parameter name from column H
            param_name = safe_get_cell_value(summary, h_cell, 'str')
            
            # Get count from column I (same row)
            row = h_cell[1:]
            i_cell = f'I{row}'
            count = safe_get_cell_value(summary, i_cell, 'int')
            
            status = "✅" if count > 0 else "  "
            if count > 0 or row in ['3', '4', '5']:  # Show first few regardless
                print(f"{h_cell:<8} {param_name or 'N/A':<20} {count:<15} {status}")
            
            if count > 0:
                discrete_found += 1
        
        print(f"\nDiscrete exceedances found: {discrete_found}/23")
        
        # Summary
        total_exceedances = continuous_found + discrete_found
        print("\n" + "=" * 80)
        print(f"SUMMARY:")
        print(f"  Total exceedances that would be inserted: {total_exceedances}")
        print(f"  - Continuous: {continuous_found}")
        print(f"  - Discrete: {discrete_found}")
        
        if total_exceedances == 0:
            print("\n⚠️  WARNING: No exceedances found!")
            print("   Possible causes:")
            print("   1. All exceedance counts are 0 (normal for a clean flight)")
            print("   2. Cell references are incorrect")
            print("   3. Data format issues")
            
            # Show raw values from a few cells for debugging
            print("\n🔍 Raw Cell Values (for debugging):")
            test_cells = ['C9', 'C10', 'C11', 'I3', 'I4', 'I5']
            for cell in test_cells:
                raw_value = summary[cell].value
                print(f"   {cell}: {raw_value} (type: {type(raw_value).__name__})")
        
        print("=" * 80)
        
        workbook.close()
        
    except Exception as e:
        print(f"❌ ERROR: {e}")
        import traceback
        traceback.print_exc()

# Main execution
if __name__ == "__main__":
    import sys
    
    if len(sys.argv) > 1:
        file_path = sys.argv[1]
    else:
        # Default test file
        file_path = r"A:\populate_fdap_db\flight_data\UNO-561P_01-10-25_1.xlsm"
    
    diagnose_exceedances(file_path)
