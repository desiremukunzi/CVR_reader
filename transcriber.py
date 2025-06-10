import whisper
import pandas as pd
from rapidfuzz import fuzz
from openpyxl import load_workbook

EXCEL_FILE = "workbench_cvr.xlsm"  # Your Excel file
SHEET_NAME = "Checklist"  # The sheet containing the checklist
CHECKED_COLUMN = "B"  # Column where check results will be written (2nd column)
THRESHOLD = 70  # Minimum similarity percentage for checkeding

# Load checklist from Excel
def load_checklist_from_excel(filename=EXCEL_FILE, sheet_name=SHEET_NAME):
    df = pd.read_excel(filename, sheet_name=sheet_name, engine="openpyxl")
    checklist = df.iloc[:, 0].dropna().tolist()  # Read first column (checklist)
    return df, checklist

# Check compliance with fuzzy matching
def check_compliance(transcript, checklist):
    transcript_lower = transcript.lower()
    results = []

    for step in checklist:
        highest_score = max(
            fuzz.partial_ratio(step.lower(), sentence.lower()) 
            for sentence in transcript_lower.split('.')
        )
        if highest_score >= THRESHOLD:
            results.append(("checked", step, highest_score))  # checked
        else:
            results.append(("skipped", step, highest_score))  # skipped

    return results

# Transcribe CVR audio
def transcribe_audio(audio_file):
    model = whisper.load_model("medium")
    result = model.transcribe(audio_file,language="en")
    return result["text"]

# Update Excel with results
def update_excel_with_results(results, filename=EXCEL_FILE, sheet_name=SHEET_NAME):
    wb = load_workbook(filename,keep_vba=True)
    wb.save('sheet_name')
    ws = wb[sheet_name]
    wb.close()



    row = 2  # Checks start from row 2
    for result in results:
        ws[f"{CHECKED_COLUMN}{row}"] = result[0]  # "checked" or "skipped"
        row += 1

    wb.save(filename)
    print(f"✅ Results updated in {filename} successfully!")

# Main execution
if __name__ == "__main__":
    df, checklist = load_checklist_from_excel()
    transcript = transcribe_audio("RAF1510-03-04-25.wav")

    print("\n--- Transcript ---")
    print(transcript)

    results = check_compliance(transcript, checklist)

    print("\n--- Checklist Compliance Report ---")
    for result in results:
        print(f"{result[0]} {result[1]} (Match: {result[2]}%)")

    update_excel_with_results(results)
