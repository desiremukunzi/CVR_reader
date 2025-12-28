"""
Flight Data Analysis and Processing System (FDAPS)
CORRECTED VERSION - Matches actual database schema
"""

import os
import sys
import logging
from datetime import datetime, date
from pathlib import Path
from typing import Dict, List, Optional, Tuple
import re

import mysql.connector
from mysql.connector import Error, pooling
import openpyxl
from openpyxl import load_workbook
from dotenv import load_dotenv

# Import FlightAnalyzer from same directory
try:
    from flight_analyzer import FlightAnalyzer
    ANOMALY_DETECTION_AVAILABLE = True
    logger_msg = "FlightAnalyzer imported successfully"
except ImportError as e:
    ANOMALY_DETECTION_AVAILABLE = False
    logger_msg = f"Could not import FlightAnalyzer: {e}"
    class FlightAnalyzer:
        def __init__(self, data_folder='flight_data'):
            pass
        def analyze_flight(self, file_path, sheet_name='Clean Data'):
            return {'anomalies': []}

# ============================================================================
# CONFIGURATION
# ============================================================================

load_dotenv()

FOLDER_PATH = r"A:\Onedrive\RAF-61504\August\NORMAL"
START_DATE = date(2025, 1, 1)
END_DATE = date(2025, 12, 31)
AIRCRAFT_ID = 2

# Exceedance parameters mapping
EXCEEDANCE_PARAMS = {
    'B9': 'IAS', 'B10': 'Alt', 'B11': 'Roll', 'B12': 'PITCH',
    'B13': 'Fcp', 'B14': 'N1/N2 Split', 'B15': 'N1', 'B16': 'N2', 'B17': 'Nmr',
    'H3': 'iAPr/p', 'H4': 'iChips', 'H5': 'iEMG1', 'H6': 'iEMG2',
    'H7': 'iF_gen1', 'H8': 'iF_gen2', 'H9': 'iF_pump1', 'H10': 'iF_pump2',
    'H11': 'iF_pumpS', 'H12': 'iFire_KO-50', 'H13': 'iFire_mgb', 'H14': 'iFire_v1',
    'H15': 'iFire_v2', 'H16': 'iFire1', 'H17': 'iFire2', 'H18': 'iHSaux',
    'H19': 'iHSmain', 'H20': 'inFT1', 'H21': 'inFT2', 'H22': 'iOP_mgb',
    'H23': 'iOP1', 'H24': 'iOP2', 'H25': 'iQTmin',
}

CHECKLIST_SHEETS = {
    'STARTING WITH AC-GPU CHECKLIST': 1,
    'STARTING WITH DC-GPU CHECKLIST': 2,
    'STARTING WITHOUT GPU CHECKLIST': 3,
}

CHECKLIST_PROCESSING_ORDER = [
    'STARTING WITH AC-GPU CHECKLIST',
    'STARTING WITHOUT GPU CHECKLIST',
    'STARTING WITH DC-GPU CHECKLIST',
]

MAX_CHECKLIST_ROW = 180
ENABLE_ANOMALY_DETECTION = ANOMALY_DETECTION_AVAILABLE
FLIGHT_ANALYZER_DATA_FOLDER = 'flight_data'

PHASE_NAME_TO_ID = {
    'before takeoff': 1,
    'when airborne': 2,
    'after landing': 3,
    'unknown': 1
}

# ============================================================================
# LOGGING SETUP
# ============================================================================

def setup_logging():
    log_dir = Path("logs")
    log_dir.mkdir(exist_ok=True)
    log_filename = log_dir / f"fdaps_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(levelname)s - %(message)s',
        handlers=[
            logging.FileHandler(log_filename, encoding='utf-8'),
            logging.StreamHandler(sys.stdout)
        ]
    )
    return logging.getLogger(__name__)

logger = setup_logging()
logger.info(logger_msg)

# ============================================================================
# DATABASE CONNECTION
# ============================================================================

class DatabaseManager:
    def __init__(self):
        self.connection_pool = None
        self.initialize_pool()
    
    def initialize_pool(self):
        try:
            self.connection_pool = pooling.MySQLConnectionPool(
                pool_name="fdaps_pool",
                pool_size=5,
                pool_reset_session=True,
                host=os.getenv('DB_HOST', 'localhost'),
                port=int(os.getenv('DB_PORT', 3306)),
                database='fdap_new',
                user=os.getenv('DB_USER', 'root'),
                password=os.getenv('DB_PASSWORD', ''),
                charset='utf8mb4',
                collation='utf8mb4_general_ci'
            )
            logger.info("Database connection pool initialized successfully")
        except Error as e:
            logger.error(f"Error initializing database pool: {e}")
            raise
    
    def get_connection(self):
        try:
            return self.connection_pool.get_connection()
        except Error as e:
            logger.error(f"Error getting connection from pool: {e}")
            raise
    
    def insert_or_update_flight(self, flight_data: Dict) -> Optional[int]:
        """
        Insert or update flight record
        UPDATED: Includes anomalies and anomalies_percentage
        """
        connection = None
        cursor = None
        
        try:
            connection = self.get_connection()
            cursor = connection.cursor()
            
            check_query = """
                SELECT id FROM flights 
                WHERE flight_date = %s AND PIC = %s AND SIC = %s AND FE = %s AND sortie = %s
            """
            cursor.execute(check_query, (
                flight_data['flight_date'],
                flight_data['PIC'],
                flight_data['SIC'],
                flight_data['FE'],
                flight_data['sortie']
            ))
            
            existing = cursor.fetchone()
            
            if existing:
                flight_id = existing[0]
                update_query = """
                    UPDATE flights 
                    SET checks_not_complied = %s,
                        compliance_percentage = %s,
                        continuous_exceedances = %s,
                        discrete_exceedances = %s,
                        anomalies = %s,
                        anomalies_percentage = %s,
                        updated_at = CURRENT_TIMESTAMP
                    WHERE id = %s
                """
                cursor.execute(update_query, (
                    flight_data['checks_not_complied'],
                    flight_data['compliance_percentage'],
                    flight_data.get('continuous_exceedances', 0),
                    flight_data.get('discrete_exceedances', 0),
                    flight_data.get('anomalies', 0),
                    flight_data.get('anomalies_percentage', 0.0),
                    flight_id
                ))
                logger.info(f"Updated flight ID {flight_id}")
            else:
                insert_query = """
                    INSERT INTO flights 
                    (flight_date, aircraft_id, PIC, SIC, FE, flight_type_id, sortie, 
                     checks_not_complied, compliance_percentage, continuous_exceedances, 
                     discrete_exceedances, anomalies, anomalies_percentage)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                """
                cursor.execute(insert_query, (
                    flight_data['flight_date'],
                    flight_data['aircraft_id'],
                    flight_data['PIC'],
                    flight_data['SIC'],
                    flight_data['FE'],
                    flight_data['flight_type_id'],
                    flight_data['sortie'],
                    flight_data['checks_not_complied'],
                    flight_data['compliance_percentage'],
                    flight_data.get('continuous_exceedances', 0),
                    flight_data.get('discrete_exceedances', 0),
                    flight_data.get('anomalies', 0),
                    flight_data.get('anomalies_percentage', 0.0)
                ))
                flight_id = cursor.lastrowid
                logger.info(f"Inserted new flight ID {flight_id}")
            
            connection.commit()
            return flight_id
            
        except Error as e:
            if connection:
                connection.rollback()
            logger.error(f"Database error in insert_or_update_flight: {e}")
            return None
        finally:
            if cursor:
                cursor.close()
            if connection:
                connection.close()
    
    def delete_flight_exceedances(self, flight_id: int):
        connection = None
        cursor = None
        try:
            connection = self.get_connection()
            cursor = connection.cursor()
            cursor.execute("DELETE FROM exceedances WHERE flight_id = %s", (flight_id,))
            connection.commit()
        except Error as e:
            if connection:
                connection.rollback()
            logger.error(f"Error deleting exceedances: {e}")
        finally:
            if cursor:
                cursor.close()
            if connection:
                connection.close()
    
    def insert_exceedances(self, flight_id: int, exceedances: List[Dict]) -> bool:
        if not exceedances:
            return True
        
        connection = None
        cursor = None
        try:
            connection = self.get_connection()
            cursor = connection.cursor()
            
            insert_query = """
                INSERT INTO exceedances 
                (flight_id, parameter_MI_17V_5_name, number_of_exceedances)
                VALUES (%s, %s, %s)
            """
            
            exceedance_data = [
                (flight_id, exc['parameter'], exc['count'])
                for exc in exceedances
            ]
            
            cursor.executemany(insert_query, exceedance_data)
            connection.commit()
            logger.info(f"Inserted {len(exceedances)} exceedance records")
            return True
        except Error as e:
            if connection:
                connection.rollback()
            logger.error(f"Error inserting exceedances: {e}")
            return False
        finally:
            if cursor:
                cursor.close()
            if connection:
                connection.close()
    
    def delete_flight_missed_checks(self, flight_id: int):
        connection = None
        cursor = None
        try:
            connection = self.get_connection()
            cursor = connection.cursor()
            cursor.execute("DELETE FROM missed_checks WHERE flight_id = %s", (flight_id,))
            connection.commit()
        except Error as e:
            if connection:
                connection.rollback()
            logger.error(f"Error deleting missed checks: {e}")
        finally:
            if cursor:
                cursor.close()
            if connection:
                connection.close()
    
    def insert_missed_checks(self, flight_id: int, missed_checks: List[Dict]) -> bool:
        if not missed_checks:
            return True
        
        connection = None
        cursor = None
        try:
            connection = self.get_connection()
            cursor = connection.cursor()
            
            insert_query = """
                INSERT INTO missed_checks 
                (flight_id, checklist_type_id, checklist_item_position)
                VALUES (%s, %s, %s)
            """
            
            check_data = [
                (flight_id, check['checklist_type_id'], check['position'])
                for check in missed_checks
            ]
            
            cursor.executemany(insert_query, check_data)
            connection.commit()
            logger.info(f"Inserted {len(missed_checks)} missed check records")
            return True
        except Error as e:
            if connection:
                connection.rollback()
            logger.error(f"Error inserting missed checks: {e}")
            return False
        finally:
            if cursor:
                cursor.close()
            if connection:
                connection.close()
    
    def delete_flight_anomalies(self, flight_id: int):
        connection = None
        cursor = None
        try:
            connection = self.get_connection()
            cursor = connection.cursor()
            cursor.execute("DELETE FROM anomalies WHERE flight_id = %s", (flight_id,))
            connection.commit()
        except Error as e:
            if connection:
                connection.rollback()
            logger.error(f"Error deleting anomalies: {e}")
        finally:
            if cursor:
                cursor.close()
            if connection:
                connection.close()
    
    def insert_anomalies(self, flight_id: int, anomalies: List[Dict]) -> bool:
        """
        Insert anomaly records
        CORRECTED: Uses total_anomalies (int) instead of anomaly_score/description
        """
        if not anomalies:
            return True
        
        connection = None
        cursor = None
        try:
            connection = self.get_connection()
            cursor = connection.cursor()
            
            insert_query = """
                INSERT INTO anomalies 
                (flight_id, parameter_MI_17V_5_name, phase_of_flight_id, total_anomalies)
                VALUES (%s, %s, %s, %s)
            """
            
            anomaly_data = [
                (
                    flight_id,
                    anom['parameter'],
                    anom['phase_of_flight_id'],
                    anom['total_anomalies']
                )
                for anom in anomalies
            ]
            
            cursor.executemany(insert_query, anomaly_data)
            connection.commit()
            logger.info(f"Inserted {len(anomalies)} anomaly records")
            return True
        except Error as e:
            if connection:
                connection.rollback()
            logger.error(f"Error inserting anomalies: {e}")
            return False
        finally:
            if cursor:
                cursor.close()
            if connection:
                connection.close()

# ============================================================================
# EXCEL PROCESSING
# ============================================================================

class FlightDataExtractor:
    
    @staticmethod
    def safe_get_cell_value(sheet, cell_ref, value_type='int'):
        try:
            value = sheet[cell_ref].value
            if value is None or value == '' or value == '-':
                return 0 if value_type in ['int', 'float'] else None
            if value_type == 'int':
                return int(float(value))
            elif value_type == 'float':
                return float(value)
            elif value_type == 'str':
                return str(value).strip()
            else:
                return value
        except (ValueError, TypeError, AttributeError) as e:
            logger.debug(f"Error getting value from {cell_ref}: {e}")
            return 0 if value_type in ['int', 'float'] else None
    
    def extract_flight_info(self, workbook, flight_date: date, sortie: int) -> Optional[Dict]:
        try:
            if 'Summary' not in workbook.sheetnames:
                return None
            
            summary = workbook['Summary']
            
            pic = self.safe_get_cell_value(summary, 'B2', 'str')
            sic = self.safe_get_cell_value(summary, 'B3', 'str')
            fe = self.safe_get_cell_value(summary, 'B4', 'str')
            
            if not all([pic, sic, fe]):
                return None
            
            checks_not_complied = self.safe_get_cell_value(summary, 'F9', 'int')
            value = self.safe_get_cell_value(summary, 'F10', 'str') or '0'
            compliance_percentage = float(value.replace('%', ''))
            continuous_exceedances = self.safe_get_cell_value(summary, 'B18', 'int')
            
            discrete_exceedances = 0
            for row in range(3, 26):
                cell_value = self.safe_get_cell_value(summary, f'H{row}', 'int')
                discrete_exceedances += cell_value
            
            flight_data = {
                'flight_date': flight_date,
                'aircraft_id': AIRCRAFT_ID,
                'PIC': pic,
                'SIC': sic,
                'FE': fe,
                'flight_type_id': 1,
                'sortie': sortie,
                'checks_not_complied': checks_not_complied,
                'compliance_percentage': compliance_percentage,
                'continuous_exceedances': continuous_exceedances,
                'discrete_exceedances': discrete_exceedances,
                # Anomaly counts will be added later
                'anomalies': 0,
                'anomalies_percentage': 0.0
            }
            
            return flight_data
            
        except Exception as e:
            logger.error(f"Error extracting flight info: {e}")
            return None
    
    def extract_exceedances(self, workbook) -> List[Dict]:
        exceedances = []
        try:
            if 'Summary' not in workbook.sheetnames:
                return exceedances
            
            summary = workbook['Summary']
            
            continuous_params = {
                'B9': 'IAS', 'B10': 'Alt', 'B11': 'Roll', 'B12': 'PITCH',
                'B13': 'Fcp', 'B14': 'N1/N2 Split', 'B15': 'N1', 'B16': 'N2', 'B17': 'Nmr',
            }
            
            for cell_ref, param_name in continuous_params.items():
                count = self.safe_get_cell_value(summary, cell_ref, 'int')
                if count > 0:
                    exceedances.append({'parameter': param_name, 'count': count})
            
            discrete_params = {
                'H3': 'iAPr/p', 'H4': 'iChips', 'H5': 'iEMG1', 'H6': 'iEMG2',
                'H7': 'iF_gen1', 'H8': 'iF_gen2', 'H9': 'iF_pump1', 'H10': 'iF_pump2',
                'H11': 'iF_pumpS', 'H12': 'iFire_KO-50', 'H13': 'iFire_mgb', 'H14': 'iFire_v1',
                'H15': 'iFire_v2', 'H16': 'iFire1', 'H17': 'iFire2', 'H18': 'iHSaux',
                'H19': 'iHSmain', 'H20': 'inFT1', 'H21': 'inFT2', 'H22': 'iOP_mgb',
                'H23': 'iOP1', 'H24': 'iOP2', 'H25': 'iQTmin',
            }
            
            for cell_ref, param_name in discrete_params.items():
                count = self.safe_get_cell_value(summary, cell_ref, 'int')
                if count > 0:
                    exceedances.append({'parameter': param_name, 'count': count})
            
        except Exception as e:
            logger.error(f"Error extracting exceedances: {e}")
        
        return exceedances
    
    def extract_missed_checks(self, workbook) -> List[Dict]:
        missed_checks = []
        try:
            for sheet_name in CHECKLIST_PROCESSING_ORDER:
                if sheet_name not in workbook.sheetnames:
                    continue
                
                sheet = workbook[sheet_name]
                checklist_type_id = CHECKLIST_SHEETS[sheet_name]
                has_marks = False
                current_missed_checks = []
                
                for row in range(2, MAX_CHECKLIST_ROW + 1):
                    cell_value = sheet[f'B{row}'].value
                    if cell_value is not None:
                        cell_str = str(cell_value).strip().upper()
                        if '✔' in str(cell_value) or cell_str in ['✓', 'V', 'OK']:
                            has_marks = True
                            continue
                        if cell_str == '✘':
                            has_marks = True
                            current_missed_checks.append({
                                'checklist_type_id': checklist_type_id,
                                'position': row
                            })
                
                if has_marks:
                    missed_checks = current_missed_checks
                    break
        except Exception as e:
            logger.error(f"Error extracting missed checks: {e}")
        
        return missed_checks
    
    def extract_anomalies(self, workbook, file_path: Path) -> Tuple[List[Dict], int, float]:
        """
        Extract anomalies using FlightAnalyzer
        CORRECTED: Returns grouped anomalies with total_anomalies count
        
        Returns:
            Tuple of (anomalies_list, total_count, percentage)
        """
        anomalies = []
        total_anomaly_count = 0
        anomaly_percentage = 0.0
        
        if not ENABLE_ANOMALY_DETECTION:
            return anomalies, total_anomaly_count, anomaly_percentage
        
        try:
            if 'Clean Data' not in workbook.sheetnames:
                logger.warning("Clean Data sheet not found")
                return anomalies, total_anomaly_count, anomaly_percentage
            
            analyzer = FlightAnalyzer(data_folder=FLIGHT_ANALYZER_DATA_FOLDER)
            logger.info(f"Running anomaly detection on {file_path.name}")
            results = analyzer.analyze_flight(str(file_path), sheet_name='Clean Data')
            
            if results and 'anomalies' in results:
                raw_anomalies = results['anomalies']
                total_anomaly_count = len(raw_anomalies)
                
                # Group anomalies by parameter and phase
                anomaly_groups = {}
                for anom in raw_anomalies:
                    phase_name = anom.get('phase', 'unknown')
                    phase_id = PHASE_NAME_TO_ID.get(phase_name, 1)
                    param = anom.get('parameter', 'Unknown')
                    
                    key = (param, phase_id)
                    if key not in anomaly_groups:
                        anomaly_groups[key] = 0
                    anomaly_groups[key] += 1
                
                # Create anomaly records
                for (param, phase_id), count in anomaly_groups.items():
                    anomalies.append({
                        'parameter': param,
                        'phase_of_flight_id': phase_id,
                        'total_anomalies': count
                    })
                
                # Calculate percentage
                total_data_points = results.get('total_data_points', 0)
                if total_data_points > 0:
                    anomaly_percentage = round((total_anomaly_count / total_data_points) * 100, 2)
                
                logger.info(f"Detected {total_anomaly_count} total anomalies ({len(anomalies)} unique param/phase combinations)")
            
        except Exception as e:
            logger.error(f"Error detecting anomalies: {e}", exc_info=True)
        
        return anomalies, total_anomaly_count, anomaly_percentage

# ============================================================================
# FILE PROCESSING
# ============================================================================

class FlightFileProcessor:
    
    def __init__(self, db_manager: DatabaseManager):
        self.db_manager = db_manager
        self.extractor = FlightDataExtractor()
    
    def parse_filename(self, filename: str) -> Optional[Tuple[date, int]]:
        try:
            name_without_ext = filename.rsplit('.', 1)[0]
            parts = name_without_ext.split('_')
            if len(parts) < 3:
                return None
            
            date_str = parts[1]
            date_parts = date_str.split('-')
            if len(date_parts) != 3:
                return None
            
            day, month, year = int(date_parts[0]), int(date_parts[1]), int(date_parts[2])
            if year < 100:
                year = 2000 + year
            
            flight_date = date(year, month, day)
            sortie = int(parts[2])
            return (flight_date, sortie)
        except (ValueError, IndexError) as e:
            logger.debug(f"Error parsing filename {filename}: {e}")
            return None
    
    def process_file(self, file_path: Path) -> bool:
        try:
            logger.info(f"Processing file: {file_path.name}")
            
            parsed = self.parse_filename(file_path.name)
            if not parsed:
                logger.warning(f"Could not parse filename: {file_path.name}")
                return False
            
            flight_date, sortie = parsed
            
            if not (START_DATE <= flight_date <= END_DATE):
                logger.debug(f"File date {flight_date} outside range")
                return False
            
            try:
                workbook = load_workbook(file_path, read_only=True, data_only=True)
            except Exception as e:
                logger.error(f"Error loading workbook: {e}")
                return False
            
            flight_data = self.extractor.extract_flight_info(workbook, flight_date, sortie)
            if not flight_data:
                workbook.close()
                return False
            
            # Extract anomalies FIRST to get counts
            anomalies, total_anomaly_count, anomaly_percentage = self.extractor.extract_anomalies(workbook, file_path)
            
            # Add anomaly counts to flight_data
            flight_data['anomalies'] = total_anomaly_count
            flight_data['anomalies_percentage'] = anomaly_percentage
            
            # Insert/update flight
            flight_id = self.db_manager.insert_or_update_flight(flight_data)
            if not flight_id:
                workbook.close()
                return False
            
            # Delete existing data
            self.db_manager.delete_flight_exceedances(flight_id)
            self.db_manager.delete_flight_missed_checks(flight_id)
            self.db_manager.delete_flight_anomalies(flight_id)
            
            # Insert new data
            exceedances = self.extractor.extract_exceedances(workbook)
            self.db_manager.insert_exceedances(flight_id, exceedances)
            
            missed_checks = self.extractor.extract_missed_checks(workbook)
            self.db_manager.insert_missed_checks(flight_id, missed_checks)
            
            self.db_manager.insert_anomalies(flight_id, anomalies)
            
            workbook.close()
            logger.info(f"Successfully processed {file_path.name}")
            return True
            
        except Exception as e:
            logger.error(f"Unexpected error: {e}", exc_info=True)
            return False

# ============================================================================
# MAIN EXECUTION
# ============================================================================

def main():
    logger.info("=" * 80)
    logger.info("Flight Data Analysis and Processing System (FDAPS)")
    logger.info("=" * 80)
    logger.info(f"Folder: {FOLDER_PATH}")
    logger.info(f"Date Range: {START_DATE} to {END_DATE}")
    logger.info(f"Anomaly Detection: {'Enabled' if ENABLE_ANOMALY_DETECTION else 'Disabled'}")
    logger.info("=" * 80)
    
    folder = Path(FOLDER_PATH)
    if not folder.exists():
        logger.error(f"Folder not found: {FOLDER_PATH}")
        return
    
    try:
        db_manager = DatabaseManager()
    except Exception as e:
        logger.error(f"Failed to initialize database: {e}")
        return
    
    processor = FlightFileProcessor(db_manager)
    excel_files = list(folder.glob("*.xls*"))
    logger.info(f"Found {len(excel_files)} Excel files")
    
    if not excel_files:
        logger.warning("No Excel files found")
        return
    
    successful = failed = skipped = 0
    
    for file_path in excel_files:
        result = processor.process_file(file_path)
        if result:
            successful += 1
        else:
            failed += 1
    
    logger.info("=" * 80)
    logger.info("Processing Summary:")
    logger.info(f"  Total: {len(excel_files)}")
    logger.info(f"  Successful: {successful}")
    logger.info(f"  Failed: {failed}")
    logger.info("=" * 80)

if __name__ == "__main__":
    main()