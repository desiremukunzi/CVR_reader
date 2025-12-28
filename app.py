import os
import pandas as pd
from flask import Flask, render_template, request, jsonify, send_from_directory, url_for, session
from werkzeug.utils import secure_filename
from rapidfuzz import fuzz
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
import subprocess
from statistics import mean
import re
from faster_whisper import WhisperModel
import tempfile
import shutil
from datetime import datetime, date

# Try to import enhanced FlightAnalyzer with database support
# Falls back to regular FlightAnalyzer if database version not available
try:
    from flight_analyzer_with_db import FlightAnalyzer
    DATABASE_ENABLED = True
    print("Using FlightAnalyzer with database support")
except ImportError:
    from populate_db import FlightAnalyzer
    DATABASE_ENABLED = False
    print("Warning: Using FlightAnalyzer without database support")


# ============================================================================
# CONFIGURATION
# ============================================================================

# Exceedance parameters mapping (Summary sheet cell location -> parameter name)
# This matches the VBA FDR analysis system output format
EXCEEDANCE_PARAMS = {
    'B9': 'IAS',
    'B10': 'Alt',
    'B11': 'Roll',
    'B12': 'PITCH',
    'B13': 'Fcp',
    'B14': 'N1/N2 Split',
    'B15': 'N1',
    'B16': 'N2',
    'B17': 'Nmr',
    'H3': 'iAPr/p',
    'H4': 'iChips',
    'H5': 'iEMG1',
    'H6': 'iEMG2',
    'H7': 'iF_gen1',
    'H8': 'iF_gen2',
    'H9': 'iF_pump1',
    'H10': 'iF_pump2',
    'H11': 'iF_pumpS',
    'H12': 'iFire_KO-50',
    'H13': 'iFire_mgb',
    'H14': 'iFire_v1',
    'H15': 'iFire_v2',
    'H16': 'iFire1',
    'H17': 'iFire2',
    'H18': 'inFT1',
    'H19': 'inFT2',
    'H20': 'iOP_mgb',
    'H21': 'iOP1',
    'H22': 'iOP2',
    'H23': 'iQTmin',
}

# Flask setup
app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'dev-secret-key-change-in-production-12345')  # NEW: Required for sessions

# Define folders
UPLOAD_FOLDER = "uploads"
COMPLIANCE_EXCEL_OUTPUT = "compliance_excel_output"
CHECKED_COLUMN = "B"
TRANSCRIPT_FOLDER = "transcripts"
COMPLIANCE_TEXT_REPORTS_FOLDER = "compliance_text_reports"
FLIGHT_DATA_FOLDER = "flight_data"

# Assign to app.config
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['COMPLIANCE_EXCEL_OUTPUT'] = COMPLIANCE_EXCEL_OUTPUT
app.config['TRANSCRIPT_FOLDER'] = TRANSCRIPT_FOLDER
app.config['COMPLIANCE_TEXT_REPORTS_FOLDER'] = COMPLIANCE_TEXT_REPORTS_FOLDER
app.config['FLIGHT_DATA_FOLDER'] = FLIGHT_DATA_FOLDER

# Ensure necessary folders exist
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
os.makedirs(app.config['COMPLIANCE_EXCEL_OUTPUT'], exist_ok=True)
os.makedirs("workbench", exist_ok=True)
os.makedirs(app.config['TRANSCRIPT_FOLDER'], exist_ok=True)
os.makedirs(app.config['COMPLIANCE_TEXT_REPORTS_FOLDER'], exist_ok=True)
os.makedirs(app.config['FLIGHT_DATA_FOLDER'], exist_ok=True)

# Initialize Whisper model
print("Initializing Whisper model...")
model = WhisperModel("medium", device="cuda", compute_type="float16")

# Initialize Flight Analyzer (UPDATED)
print("Initializing FlightAnalyzer...")
if DATABASE_ENABLED:
    # Initialize with database support
    flight_analyzer = FlightAnalyzer(
        data_folder=app.config['FLIGHT_DATA_FOLDER'],
        enable_database=True
    )
else:
    # Initialize without database
    flight_analyzer = FlightAnalyzer(data_folder=app.config['FLIGHT_DATA_FOLDER'])

# Load historical data if available
historical_folder = os.path.join(app.config['FLIGHT_DATA_FOLDER'], 'historical')
if os.path.exists(historical_folder):
    try:
        print(f"Loading historical data from: {historical_folder}")
        flight_analyzer.load_historical_from_folder(historical_folder)
        historical_count = (
            flight_analyzer.historical_data['flight_id'].nunique()
            if hasattr(flight_analyzer, 'historical_data') and not flight_analyzer.historical_data.empty
            else 0
        )
        print(f"✓ Loaded {historical_count} historical flights")
    except Exception as e:
        print(f"Note: Could not load historical data: {e}")
else:
    print(f"Note: Historical data folder not found at: {historical_folder}")

# Store the latest anomaly report (for backward compatibility)
latest_anomaly_report = None


def preprocess_audio(input_path):
    """Skip preprocessing - use original WAV files."""
    print(f"Using original WAV: {input_path}")
    return input_path


def concatenate_audio_files(input_paths, output_filename, upload_folder):
    """Concatenate multiple audio files into a single WAV file."""
    concat_list_path = os.path.join(tempfile.gettempdir(), "files.txt")

    with open(concat_list_path, "w") as f:
        for path in input_paths:
            f.write(f"file '{path}'\n")

    if not output_filename.lower().endswith(".wav"):
        output_filename += ".wav"

    output_path = os.path.join(upload_folder, secure_filename(output_filename))

    command = [
        "ffmpeg", "-y",
        "-f", "concat",
        "-safe", "0",
        "-i", concat_list_path,
        "-c", "copy",
        output_path
    ]

    try:
        result = subprocess.run(command, capture_output=True, text=True, check=True)
        print(f"FFmpeg stdout (concatenation): {result.stdout}")
        return output_path
    except subprocess.CalledProcessError as e:
        print(f"FFmpeg failed during concatenation: {e}")
        print(f"FFmpeg stderr (concatenation): {e.stderr}")
        return None
    finally:
        if os.path.exists(concat_list_path):
            os.remove(concat_list_path)


def load_checklist(excel_file_path, sheet_name):
    """
    Load checklist items from Excel sheet.
    Returns: (df, checklist_items, row_positions)
    
    row_positions maps item index to Excel row number (starting from row 2)
    """
    try:
        df = pd.read_excel(excel_file_path, sheet_name=sheet_name, engine="openpyxl")
        checklist_items = df.iloc[:, 0].dropna().tolist()
        
        # Create mapping: checklist item index -> Excel row number
        # Excel rows start at 1, header is row 1, data starts at row 2
        row_positions = {i: i + 2 for i in range(len(checklist_items))}
        
        return df, checklist_items, row_positions
    except Exception as e:
        raise Exception(f"Failed to load checklist from Excel: {e}")


def clean_text(text):
    """Clean text for fuzzy matching."""
    text = text.lower()
    text = re.sub(r"[^a-zA-Z0-9\s]", "", text)
    text = re.sub(r"\b(?:roger|copy|standby|okay|affirmative|negative|check)\b", "", text)
    return text.strip()


def check_compliance(transcript, checklist, threshold=50):
    """Check compliance using fuzzy matching with sliding window."""
    transcript_lower = transcript.lower()
    transcript_words_raw = transcript_lower.split()

    results = []
    MAX_CHUNK_WORDS = 20

    for step in checklist:
        step_clean = clean_text(step)
        best_score = 0
        best_chunk_raw = ""

        for i in range(len(transcript_words_raw)):
            for j in range(i + 1, min(i + MAX_CHUNK_WORDS + 1, len(transcript_words_raw) + 1)):
                current_chunk_words_raw = transcript_words_raw[i:j]
                current_chunk_raw = ' '.join(current_chunk_words_raw)
                current_chunk_clean = clean_text(current_chunk_raw)

                if not current_chunk_clean:
                    continue

                pr = fuzz.partial_ratio(step_clean, current_chunk_clean)
                tsr = fuzz.token_set_ratio(step_clean, current_chunk_clean)
                ratio = fuzz.ratio(step_clean, current_chunk_clean)

                score = max(pr, tsr, ratio) * 0.6 + mean([pr, tsr, ratio]) * 0.4

                if score > best_score:
                    best_score = score
                    best_chunk_raw = current_chunk_raw

        if best_score == 100.0 and step_clean not in clean_text(transcript):
            best_score = 99.0

        print(f"\n✅ Checklist Item: {step}")
        print(f"   🔍 Matched: \"{best_chunk_raw}\"")
        print(f"   🎯 Score: {best_score:.1f}%")

        results.append(("PASS" if best_score >= threshold else "FAIL", step, best_score, best_chunk_raw))

    return results


def update_excel(excel_input_path, results, sheet_name, not_complied_count, compliance_percent):
    """Update Excel file with compliance results."""
    try:
        wb = load_workbook(excel_input_path, keep_vba=True)
        
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found in the uploaded Excel file.")
        ws = wb[sheet_name]

        # Update checklist results
        row = 2
        for result in results:
            status_icon = "✔" if result[0] == "PASS" else "✘"
            cell = ws[f"{CHECKED_COLUMN}{row}"]
            cell.value = status_icon
            cell.font = Font(color="008000" if result[0] == "PASS" else "FF0000")
            row += 1

        # Update Summary sheet
        if "Summary" not in wb.sheetnames:
            summary_ws = wb.create_sheet("Summary")
        else:
            summary_ws = wb["Summary"]

        bold_font_white = Font(bold=True, color="FFFFFF")
        blue_background = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        thin_border_side = Side(style='thick')

        summary_ws['E8'].value = "Checklist Compliance"
        summary_ws['E8'].font = bold_font_white
        summary_ws['E8'].fill = blue_background
        summary_ws['E8'].alignment = Alignment(horizontal='center', vertical='center')
        summary_ws.merge_cells('E8:F8')

        summary_ws.row_dimensions[8].height = 24
        summary_ws.column_dimensions['E'].width = 20
        summary_ws.column_dimensions['F'].width = 10

        summary_ws['E9'].value = "Checks Not Complied:"
        summary_ws['E9'].font = Font(bold=True)
        summary_ws['F9'].value = not_complied_count
        summary_ws['F9'].font = Font(bold=True, color="FF0000")
        summary_ws['F9'].fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        summary_ws['F9'].alignment = Alignment(horizontal='center', vertical='center')

        summary_ws['E10'].value = "Complied Percentage:"
        summary_ws['E10'].font = Font(bold=True)
        summary_ws['F10'].value = f"{compliance_percent:.1f}%"
        summary_ws['F10'].font = Font(bold=True, color="008000")
        summary_ws['F10'].fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
        summary_ws['F10'].alignment = Alignment(horizontal='center', vertical='center')

        summary_ws['E8'].border = Border(left=thin_border_side, right=thin_border_side)
        summary_ws['E9'].border = Border(left=thin_border_side)
        summary_ws['F9'].border = Border(right=thin_border_side)
        summary_ws['E10'].border = Border(bottom=thin_border_side, left=thin_border_side)
        summary_ws['F10'].border = Border(bottom=thin_border_side, right=thin_border_side)

        base_name = os.path.splitext(os.path.basename(excel_input_path))[0]
        output_excel_filename = f"{base_name}.xlsm"
        output_excel_path = os.path.join(COMPLIANCE_EXCEL_OUTPUT, output_excel_filename)

        wb.save(output_excel_path)
        print(f"Updated Excel file saved to: {output_excel_path}")
        return output_excel_path

    except Exception as e:
        print(f"Error updating Excel file: {e}")
        raise Exception(f"Failed to update Excel file: {e}")


def transcribe_audio(audio_path, custom_name=None):
    """Transcribe audio using Whisper."""
    segments, info = model.transcribe(audio_path, language="en")
    transcript_text = " ".join([segment.text for segment in segments])

    if custom_name:
        base_filename = os.path.splitext(secure_filename(custom_name))[0]
    else:
        base_filename = os.path.splitext(os.path.basename(audio_path))[0]

    transcript_filename = f"{base_filename}.txt"
    transcript_path = os.path.join(TRANSCRIPT_FOLDER, transcript_filename)

    with open(transcript_path, "w", encoding="utf-8") as f:
        f.write(transcript_text)

    print(f"Transcript saved to: {transcript_path}")
    return transcript_text


def save_compliance_report(results, output_file_name):
    """Save compliance results to text file."""
    base_name = os.path.splitext(secure_filename(output_file_name))[0]
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    report_filename = f"{base_name}_compliance_report_{timestamp}.txt"
    report_path = os.path.join(COMPLIANCE_TEXT_REPORTS_FOLDER, report_filename)

    with open(report_path, "w", encoding="utf-8") as f:
        f.write(f"Compliance Report for: {output_file_name}\n")
        f.write(f"Generated On: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        f.write("-" * 50 + "\n\n")

        for status, checklist_item, score, matched_text in results:
            f.write(f"Status: {status}\n")
            f.write(f"Checklist Item: {checklist_item}\n")
            f.write(f"Matched Text: \"{matched_text}\"\n")
            f.write(f"Score: {score:.1f}%\n")
            f.write("-" * 20 + "\n")
    print(f"Compliance report saved to: {report_path}")


def extract_flight_metadata_from_excel(excel_path, excel_filename):
    """
    Extract flight metadata from Excel filename and Summary sheet.
    
    Filename format: CallSign_DD-MM-YY_Sortie
    Example: UNO-561P_17-10-25_1
    
    Summary sheet cells:
    - B2: PIC
    - B3: SIC
    - B4: FE
    
    Args:
        excel_path: Full path to Excel file
        excel_filename: Just the filename (for parsing)
    
    Returns:
        dict: Flight metadata
    """
    metadata = {
        'flight_date': date.today().strftime('%Y-%m-%d'),
        'pic': 'UNK',
        'sic': 'UNK',
        'fe': 'UNK',
        'sortie': 1,
        'aircraft_id': 3,
        'call_sign': 'UNK'
    }
    
    # Extract from filename: CallSign_DD-MM-YY_Sortie
    try:
        # Remove extension
        base_name = os.path.splitext(excel_filename)[0]
        parts = base_name.split('_')
        
        if len(parts) >= 3:
            # Extract call sign
            metadata['call_sign'] = parts[0]
            
            # Extract date: DD-MM-YY
            date_str = parts[1]
            date_parts = date_str.split('-')
            if len(date_parts) == 3:
                day, month, year = date_parts
                # Convert YY to YYYY (assume 2000s)
                full_year = f"20{year}" if len(year) == 2 else year
                # Create date in YYYY-MM-DD format
                metadata['flight_date'] = f"{full_year}-{month.zfill(2)}-{day.zfill(2)}"
                print(f"  Extracted date: {metadata['flight_date']} from {date_str}")
            
            # Extract sortie number
            try:
                metadata['sortie'] = int(parts[2])
                print(f"  Extracted sortie: {metadata['sortie']}")
            except (ValueError, IndexError):
                pass
    except Exception as e:
        print(f"  Warning: Could not parse filename '{excel_filename}': {e}")
    
    # Extract crew info from Summary sheet
    try:
        wb = load_workbook(excel_path, read_only=True, keep_vba=False, data_only=True)
        
        if 'Summary' in wb.sheetnames:
            ws = wb['Summary']
            
            # Read crew from specific cells
            pic_value = ws['B2'].value
            sic_value = ws['B3'].value
            fe_value = ws['B4'].value
            
            # Clean and validate
            if pic_value and str(pic_value).strip():
                metadata['pic'] = str(pic_value).strip().upper()
                print(f"  Extracted PIC: {metadata['pic']}")
            
            if sic_value and str(sic_value).strip():
                metadata['sic'] = str(sic_value).strip().upper()
                print(f"  Extracted SIC: {metadata['sic']}")
            
            if fe_value and str(fe_value).strip():
                metadata['fe'] = str(fe_value).strip().upper()
                print(f"  Extracted FE: {metadata['fe']}")
        else:
            print("  Warning: No 'Summary' sheet found in Excel file")
        
        wb.close()
    except Exception as e:
        print(f"  Warning: Could not read Summary sheet: {e}")
    
    return metadata



def extract_exceedances_from_excel(excel_path):
    """
    Extract exceedance counts from the Summary sheet of the analyzed Excel file.
    This reads the VBA FDR analysis results that are already in the Summary sheet.
    
    Args:
        excel_path: Path to the Excel file with Summary sheet
        
    Returns:
        List of dicts: [{'parameter': 'IAS', 'count': 5}, ...]
    """
    exceedances = []
    
    try:
        wb = load_workbook(excel_path, read_only=True, keep_vba=False, data_only=True)
        
        if 'Summary' not in wb.sheetnames:
            print("  ⚠ Warning: No Summary sheet found for exceedances extraction")
            return exceedances
        
        ws = wb['Summary']
        
        # Extract exceedance counts from specific cells
        for cell_ref, param_name in EXCEEDANCE_PARAMS.items():
            try:
                cell_value = ws[cell_ref].value
                
                # Convert to integer, handle None, empty, or '-' values
                if cell_value is None or cell_value == '' or cell_value == '-':
                    count = 0
                else:
                    count = int(float(cell_value))
                
                # Only add if count > 0
                if count > 0:
                    exceedances.append({
                        'parameter': param_name,
                        'count': count
                    })
                    
            except (ValueError, TypeError) as e:
                print(f"  ⚠ Warning: Could not read exceedance from {cell_ref}: {e}")
                continue
        
        wb.close()
        print(f"  ✓ Extracted {len(exceedances)} exceedances from Summary sheet")
        
    except Exception as e:
        print(f"  ✗ Error extracting exceedances: {e}")
    
    return exceedances


def extract_compliance_from_excel(excel_path):
    """
    Extract CVR compliance data from the Summary sheet.
    
    Reads:
    - F9: Checks Not Complied (integer)
    - F10: Compliance Percentage (formatted as "XX.X%")
    
    Args:
        excel_path: Path to the Excel file with Summary sheet
        
    Returns:
        dict: {
            'checks_not_complied': int,
            'compliance_percentage': float,
            'has_cvr_data': bool
        }
    """
    compliance_data = {
        'checks_not_complied': None,
        'compliance_percentage': None,
        'has_cvr_data': False
    }
    
    try:
        wb = load_workbook(excel_path, read_only=True, keep_vba=False, data_only=True)
        
        if 'Summary' not in wb.sheetnames:
            print("  ℹ️ No Summary sheet found for compliance extraction")
            return compliance_data
        
        ws = wb['Summary']
        
        # Extract checks not complied (F9)
        checks_not_complied_value = ws['F9'].value
        if checks_not_complied_value is not None and str(checks_not_complied_value).strip():
            try:
                compliance_data['checks_not_complied'] = int(float(checks_not_complied_value))
                compliance_data['has_cvr_data'] = True
            except (ValueError, TypeError):
                pass
        
        # Extract compliance percentage (F10)
        compliance_percent_value = ws['F10'].value
        if compliance_percent_value is not None and str(compliance_percent_value).strip():
            try:
                # Remove % symbol if present and convert to float
                percent_str = str(compliance_percent_value).replace('%', '').strip()
                compliance_data['compliance_percentage'] = float(percent_str)
                compliance_data['has_cvr_data'] = True
            except (ValueError, TypeError):
                pass
        
        wb.close()
        
        if compliance_data['has_cvr_data']:
            print(f"  ✓ Extracted compliance from Excel:")
            print(f"     Checks Not Complied: {compliance_data['checks_not_complied']}")
            print(f"     Compliance: {compliance_data['compliance_percentage']}%")
        else:
            print("  ℹ️ No CVR compliance data found in Excel Summary sheet")
            
    except Exception as e:
        print(f"  ⚠️ Warning: Could not extract compliance data from Excel: {e}")
    
    return compliance_data


def extract_missed_checks_from_excel(excel_path):
    """
    Extract individual missed check items from the checklist sheet.
    Reads column B to find rows with ✘ marks (failed checks).
    
    Args:
        excel_path: Path to the Excel file with checklist sheet
        
    Returns:
        dict: {
            'missed_checks': [(item, score, excel_row), ...],
            'checklist_type_id': int,
            'sheet_name': str
        }
    """
    result = {
        'missed_checks': [],
        'checklist_type_id': 1,  # Default to AC-GPU
        'sheet_name': None
    }
    
    # Checklist sheet names and their IDs
    checklist_sheets = {
        'STARTING WITH AC-GPU CHECKLIST': 1,
        'STARTING WITH DC-GPU CHECKLIST': 2,
        'STARTING WITHOUT GPU CHECKLIST': 3
    }
    
    try:
        wb = load_workbook(excel_path, read_only=True, keep_vba=False, data_only=True)
        
        # Find which checklist sheet exists and has data
        for sheet_name, checklist_id in checklist_sheets.items():
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                result['sheet_name'] = sheet_name
                result['checklist_type_id'] = checklist_id
                
                # Read column B starting from row 2 (row 1 is header)
                # The column should have ✔ or ✘ marks
                missed_checks = []
                row_num = 2
                
                while True:
                    cell_value = ws[f'B{row_num}'].value
                    
                    # Stop if we hit an empty cell
                    if cell_value is None:
                        break
                    
                    # Check if this row has a ✘ (failed check)
                    if cell_value == '✘':
                        # Read the checklist item from column A
                        checklist_item = ws[f'A{row_num}'].value
                        if checklist_item:
                            # Format: (item, score, excel_row)
                            # We don't have the score from Excel, so use 0.0
                            missed_checks.append((str(checklist_item).strip(), 0.0, row_num))
                    
                    row_num += 1
                    
                    # Safety limit to avoid infinite loop
                    if row_num > 200:
                        break
                
                result['missed_checks'] = missed_checks
                
                if missed_checks:
                    print(f"  ✓ Extracted {len(missed_checks)} missed checks from '{sheet_name}'")
                else:
                    print(f"  ℹ️ No missed checks found in '{sheet_name}'")
                
                break  # Found the sheet, stop looking
        
        wb.close()
        
    except Exception as e:
        print(f"  ⚠️ Warning: Could not extract missed checks from Excel: {e}")
    
    return result


def extract_compliance_from_excel(excel_path):
    """
    Extract CVR compliance data from the Summary sheet.
    
    Reads:
    - F9: Checks Not Complied (integer)
    - F10: Compliance Percentage (formatted as "XX.X%")
    
    Args:
        excel_path: Path to the Excel file with Summary sheet
        
    Returns:
        dict: {
            'checks_not_complied': int,
            'compliance_percentage': float,
            'has_cvr_data': bool
        }
    """
    compliance_data = {
        'checks_not_complied': None,
        'compliance_percentage': None,
        'has_cvr_data': False
    }
    
    try:
        wb = load_workbook(excel_path, read_only=True, keep_vba=False, data_only=True)
        
        if 'Summary' not in wb.sheetnames:
            print("  ℹ️ No Summary sheet found for compliance extraction")
            return compliance_data
        
        ws = wb['Summary']
        
        # Extract checks not complied (F9)
        checks_not_complied_value = ws['F9'].value
        if checks_not_complied_value is not None and str(checks_not_complied_value).strip():
            try:
                compliance_data['checks_not_complied'] = int(float(checks_not_complied_value))
                compliance_data['has_cvr_data'] = True
            except (ValueError, TypeError):
                pass
        
        # Extract compliance percentage (F10)
        compliance_percent_value = ws['F10'].value
        if compliance_percent_value is not None and str(compliance_percent_value).strip():
            try:
                # Remove % symbol if present and convert to float
                percent_str = str(compliance_percent_value).replace('%', '').strip()
                compliance_data['compliance_percentage'] = float(percent_str)
                compliance_data['has_cvr_data'] = True
            except (ValueError, TypeError):
                pass
        
        wb.close()
        
        if compliance_data['has_cvr_data']:
            print(f"  ✓ Extracted compliance from Excel:")
            print(f"     Checks Not Complied: {compliance_data['checks_not_complied']}")
            print(f"     Compliance: {compliance_data['compliance_percentage']}%")
        else:
            print("  ℹ️ No CVR compliance data found in Excel Summary sheet")
            
    except Exception as e:
        print(f"  ⚠️ Warning: Could not extract compliance data from Excel: {e}")
    
    return compliance_data


def extract_exceedances_from_excel(excel_path):
    """
    Extract exceedance counts from the Summary sheet of the analyzed Excel file.
    This reads the VBA FDR analysis results that are already in the Summary sheet.
    
    Args:
        excel_path: Path to the Excel file with Summary sheet
        
    Returns:
        List of dicts: [{'parameter': 'IAS', 'count': 5}, ...]
    """
    exceedances = []
    
    try:
        wb = load_workbook(excel_path, read_only=True, keep_vba=False, data_only=True)
        
        if 'Summary' not in wb.sheetnames:
            print("  ⚠ Warning: No Summary sheet found for exceedances extraction")
            return exceedances
        
        ws = wb['Summary']
        
        # Extract exceedance counts from specific cells
        for cell_ref, param_name in EXCEEDANCE_PARAMS.items():
            try:
                cell_value = ws[cell_ref].value
                
                # Convert to integer, handle None, empty, or '-' values
                if cell_value is None or cell_value == '' or cell_value == '-':
                    count = 0
                else:
                    count = int(float(cell_value))
                
                # Only add if count > 0
                if count > 0:
                    exceedances.append({
                        'parameter': param_name,
                        'count': count
                    })
                    
            except (ValueError, TypeError) as e:
                print(f"  ⚠ Warning: Could not read exceedance from {cell_ref}: {e}")
                continue
        
        wb.close()
        print(f"  ✓ Extracted {len(exceedances)} exceedances from Summary sheet")
        
    except Exception as e:
        print(f"  ✗ Error extracting exceedances: {e}")
    
    return exceedances


# ============================================================================
# NEW ENDPOINTS FOR DATABASE INTEGRATION
# ============================================================================

@app.route("/add_to_training", methods=["POST"])
def add_to_training():
    """
    NEW ENDPOINT: Add flight to training data and retrain models.
    Called when user clicks "Add to Training Data" button on anomaly report page.
    """
    try:
        data = request.get_json() or {}
        flight_id = data.get('flight_id')
        
        print(f"🔄 Retraining models with accumulated flight data (flight_id: {flight_id})...")
        
        # Retrain models with all accumulated data
        flight_analyzer.train_models()
        
        # Update session metadata to mark as added to training
        if 'analysis_metadata' in session:
            session['analysis_metadata']['added_to_training'] = True
            session.modified = True
        
        # Also update global variable for current view
        global latest_anomaly_report
        if latest_anomaly_report:
            latest_anomaly_report['added_to_training'] = True
        
        print("✓ Model retraining completed successfully")
        return jsonify({'success': True})
        
    except Exception as e:
        print(f"❌ Error in add_to_training: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route("/save_to_database", methods=["POST"])
def save_to_database():
    """
    NEW ENDPOINT: Save flight record, anomalies, and missed checks to MySQL database.
    Called when user clicks "Save to Database" button on anomaly report page.
    """
    try:
        if not DATABASE_ENABLED:
            return jsonify({
                'success': False, 
                'error': 'Database support not available. Install flight_analyzer_with_db.py'
            }), 400
        
        data = request.get_json()
        frontend_metadata = data.get('flight_metadata')
        anomalies_summary = data.get('anomalies_summary')
        
        if not frontend_metadata:
            return jsonify({'success': False, 'error': 'Flight metadata is required'}), 400
        
        print(f"\n{'='*60}")
        print(f"DEBUG: Frontend sent metadata:")
        print(f"  Date: {frontend_metadata.get('flight_date')}")
        print(f"  Call Sign: {frontend_metadata.get('call_sign')}")
        print(f"  PIC: {frontend_metadata.get('pic')}")
        print(f"{'='*60}\n")
        
        # Anomalies summary is optional (may be empty if no anomalies detected)
        if not anomalies_summary:
            print("  ℹ️ No anomalies summary provided - will save flight record only")
            anomalies_summary = {}  # Empty dict instead of error
        
        # Start with frontend metadata as fallback
        flight_metadata = frontend_metadata
        
        # NEW: Retrieve CVR results from session
        cvr_results = session.get('cvr_results', None)
        
        # CRITICAL: Re-extract flight metadata from Excel file to ensure current data
        excel_path = session.get('excel_path', None)
        excel_filename = session.get('excel_filename', None)
        
        if excel_path and os.path.exists(excel_path) and excel_filename:
            print(f"  📋 Re-extracting flight metadata from current Excel file...")
            print(f"     Excel: {excel_filename}")
            # Extract fresh metadata from the Excel file
            fresh_metadata = extract_flight_metadata_from_excel(excel_path, excel_filename)
            
            # Override frontend metadata with fresh data
            if fresh_metadata and fresh_metadata.get('flight_date'):
                print(f"  ✓ Using FRESH metadata from Excel:")
                print(f"     Date: {fresh_metadata.get('flight_date')} (was: {frontend_metadata.get('flight_date')})")
                print(f"     PIC: {fresh_metadata.get('pic')} (was: {frontend_metadata.get('pic')})")
                flight_metadata = fresh_metadata
                # Update session with fresh metadata
                session['flight_metadata'] = fresh_metadata
                session.modified = True
            else:
                print(f"  ⚠️ Warning: Could not extract metadata, using frontend data")
                print(f"     Fresh metadata was: {fresh_metadata}")
        else:
            print(f"  ⚠️ Warning: Excel path not in session, using frontend metadata")
            print(f"     excel_path: {excel_path}")
            print(f"     excel_filename: {excel_filename}")
        
        # NEW: Extract exceedances from Excel file
        exceedances_list = []
        
        if excel_path and os.path.exists(excel_path):
            print(f"  📊 Extracting exceedances from Excel file...")
            exceedances_list = extract_exceedances_from_excel(excel_path)
        else:
            print(f"  ⚠️ Warning: Excel path not found in session, skipping exceedances")
        
        # NEW: Extract compliance data from Excel Summary sheet
        compliance_data_from_excel = None
        missed_checks_from_excel = None
        
        if excel_path and os.path.exists(excel_path):
            print(f"  📊 Extracting compliance data from Excel file...")
            compliance_data_from_excel = extract_compliance_from_excel(excel_path)
            
            # Also extract individual missed checks from the checklist sheet
            print(f"  📊 Extracting missed checks from Excel checklist...")
            missed_checks_from_excel = extract_missed_checks_from_excel(excel_path)
            
            # If we found compliance data in Excel, update or create cvr_results
            if compliance_data_from_excel and compliance_data_from_excel['has_cvr_data']:
                if not cvr_results:
                    # Create cvr_results from Excel data
                    cvr_results = {
                        'compliance_percent': compliance_data_from_excel['compliance_percentage'],
                        'not_complied_count': compliance_data_from_excel['checks_not_complied'],
                        'results': [],  # Will be populated below
                        'checklist_type_id': missed_checks_from_excel.get('checklist_type_id', 1) if missed_checks_from_excel else 1,
                        'sheet_name': missed_checks_from_excel.get('sheet_name', 'Unknown') if missed_checks_from_excel else 'Unknown'
                    }
                    
                    # DEBUG: Show what we're creating
                    print(f"  🔍 DEBUG: Created cvr_results with:")
                    print(f"     compliance_percent: {cvr_results['compliance_percent']}")
                    print(f"     not_complied_count: {cvr_results['not_complied_count']}")
                    print(f"     checklist_type_id: {cvr_results['checklist_type_id']}")
                    
                    # Add missed checks if available
                    if missed_checks_from_excel and missed_checks_from_excel['missed_checks']:
                        # Convert to the format expected by the database
                        # Format: (status, item, score, matched_text, excel_row)
                        cvr_results['results'] = [
                            ('FAIL', item, score, '', excel_row)  # matched_text empty as we don't have it
                            for item, score, excel_row in missed_checks_from_excel['missed_checks']
                        ]
                        print(f"  ✓ Added {len(cvr_results['results'])} missed checks to cvr_results")
                    
                    print(f"  ✓ Created cvr_results from Excel data")
                else:
                    # Update existing cvr_results with Excel data (Excel is source of truth)
                    cvr_results['compliance_percent'] = compliance_data_from_excel['compliance_percentage']
                    cvr_results['not_complied_count'] = compliance_data_from_excel['checks_not_complied']
                    
                    # DEBUG: Show what we're updating
                    print(f"  🔍 DEBUG: Updated cvr_results with:")
                    print(f"     compliance_percent: {cvr_results['compliance_percent']}")
                    print(f"     not_complied_count: {cvr_results['not_complied_count']}")
                    
                    # Update missed checks if available from Excel
                    if missed_checks_from_excel and missed_checks_from_excel['missed_checks']:
                        cvr_results['results'] = [
                            ('FAIL', item, score, '', excel_row)
                            for item, score, excel_row in missed_checks_from_excel['missed_checks']
                        ]
                        cvr_results['checklist_type_id'] = missed_checks_from_excel.get('checklist_type_id', 1)
                        cvr_results['sheet_name'] = missed_checks_from_excel.get('sheet_name', 'Unknown')
                        print(f"  ✓ Updated cvr_results with {len(cvr_results['results'])} missed checks from Excel")
                    
                    print(f"  ✓ Updated cvr_results with Excel compliance data")
        
        print(f"💾 Saving to database: Flight {flight_metadata.get('call_sign', 'N/A')}")
        if cvr_results:
            print(f"  - Including CVR results: {cvr_results['not_complied_count']} missed checks")
        if exceedances_list:
            print(f"  - Including exceedances: {len(exceedances_list)} parameters exceeded")
        
        # Convert anomalies_summary from dict format back to tuple keys
        # Frontend sends: {"Fcp_when airborne": 12, "IAS_before takeoff": 5, ...}
        # Backend needs: {("Fcp", "when airborne"): 12, ("IAS", "before takeoff"): 5, ...}
        anomalies_dict = {}
        for key, value in anomalies_summary.items():
            # Split by last underscore to separate param from phase
            # Handle phases with underscores like "when_airborne"
            parts = key.rsplit('_', 1) if '_' in key else [key, 'unknown']
            if len(parts) == 2:
                param, phase = parts
                # Replace underscores in phase back to spaces
                phase = phase.replace('_', ' ')
                anomalies_dict[(param, phase)] = value
            else:
                print(f"Warning: Could not parse anomaly key: {key}")
        
        # Convert flight_date from string to date object if needed
        if isinstance(flight_metadata.get('flight_date'), str):
            try:
                flight_metadata['flight_date'] = datetime.strptime(
                    flight_metadata['flight_date'], '%Y-%m-%d'
                ).date()
            except ValueError:
                # Try alternative format
                flight_metadata['flight_date'] = datetime.strptime(
                    flight_metadata['flight_date'], '%Y/%m/%d'
                ).date()
        
        # Ensure required fields have defaults
        flight_metadata.setdefault('sortie', 1)
        flight_metadata.setdefault('aircraft_id', 3)
        
        # Save to database using FlightAnalyzer method (now includes CVR results and exceedances)
        success = flight_analyzer._save_to_database(
            flight_metadata,
            anomalies_dict,
            cvr_results,      # CVR results with missed checks
            exceedances_list  # NEW: Exceedances from Summary sheet
        )
        
        if success:
            # Update session metadata to mark as saved to database
            if 'analysis_metadata' in session:
                session['analysis_metadata']['saved_to_database'] = True
                session.modified = True
            
            # Also update global variable for current view
            global latest_anomaly_report
            if latest_anomaly_report:
                latest_anomaly_report['saved_to_database'] = True
            
            print(f"✓ Successfully saved to database")
            return jsonify({
                'success': True,
                'flight_id': flight_metadata.get('flight_id')
            })
        else:
            return jsonify({'success': False, 'error': 'Database save operation returned False'}), 500
            
    except Exception as e:
        print(f"❌ Error in save_to_database: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/download_updated_excel/<filename>', methods=['GET'])
def download_updated_excel(filename):
    """
    NEW ENDPOINT: Download the updated Excel file with compliance results.
    """
    try:
        secure_filename_download = secure_filename(filename)
        return send_from_directory(
            directory=app.config['COMPLIANCE_EXCEL_OUTPUT'],
            path=secure_filename_download,
            as_attachment=True
        )
    except Exception as e:
        print(f"Error downloading file: {e}")
        return jsonify({'error': str(e)}), 404


# ============================================================================
# UPDATED ENDPOINTS
# ============================================================================

@app.route("/analyze_flight_anomalies", methods=["POST"])
def analyze_flight_anomalies():
    """
    UPDATED: Analyze flight data for anomalies using the "Clean Data" sheet.
    NO LONGER accepts add_to_training parameter - user decides on report page.
    """
    global latest_anomaly_report
    
    try:
        data = request.get_json()
        excel_filename = data.get('excel_filename')
        sheet_name = data.get('sheet_name', 'Clean Data')
        
        # REMOVED: add_to_training parameter - user decides later on report page
        
        if not excel_filename:
            return jsonify({'error': 'No Excel filename provided'}), 400
        
        # The Excel file is in COMPLIANCE_EXCEL_OUTPUT from the compliance check
        excel_path = os.path.join(COMPLIANCE_EXCEL_OUTPUT, excel_filename)
        
        if not os.path.exists(excel_path):
            return jsonify({'error': f'Excel file not found: {excel_filename}'}), 404
        
        print(f"\n{'='*60}")
        print(f"📊 ANALYZING FLIGHT DATA")
        print(f"{'='*60}")
        print(f"Excel file: {excel_path}")
        print(f"Sheet name: {sheet_name}")
        
        # IMPORTANT: Re-extract flight metadata from the Excel file to ensure current data
        # The Excel file now has updated Summary sheet with compliance data
        print("\\n📋 Re-extracting flight metadata from Excel file...")
        flight_metadata = extract_flight_metadata_from_excel(excel_path, excel_filename)
        
        # Override with session data if present (user may have manually entered data)
        session_metadata = session.get('flight_metadata', None)
        if session_metadata:
            print("  Merging with session metadata (preserving manual entries)...")
            # Preserve manual entries from session, but use Excel data as base
            for key in ['pic', 'sic', 'fe', 'sortie', 'aircraft_id', 'flight_date']:
                if key in session_metadata and session_metadata[key] not in ['UNK', None, '']:
                    # Only override if session has valid data
                    if key == 'flight_date':
                        # Ensure consistent format
                        if isinstance(session_metadata[key], str):
                            flight_metadata[key] = session_metadata[key]
                    else:
                        flight_metadata[key] = session_metadata[key]
        
        # Update session with fresh metadata
        session['flight_metadata'] = flight_metadata
        session.modified = True
        
        print(f"✓ Current flight metadata: {flight_metadata}")
        
        # Fallback defaults if extraction failed completely
        if not flight_metadata or flight_metadata.get('pic') == 'UNK':
            # Try to extract info from filename or use defaults
            print("⚠ Warning: No flight metadata found in session, using defaults")
            flight_metadata = {
                'flight_date': date.today(),
                'pic': 'UNK',
                'sic': 'UNK',
                'fe': 'UNK',
                'sortie': 1,
                'aircraft_id': 3,
                'call_sign': 'UNK'
            }
        else:
            # Ensure call_sign exists (might be from old session)
            flight_metadata.setdefault('call_sign', 'UNK')
            print(f"Flight metadata: {flight_metadata}")
        
        # Analyze the flight WITHOUT auto-actions (UPDATED)
        if DATABASE_ENABLED:
            results = flight_analyzer.analyze_flight(
                excel_path=excel_path,
                sheet_name=sheet_name,
                flight_metadata=flight_metadata,
                interactive=False,
                auto_add_to_training=False,  # NEW: Don't auto-add to training
                auto_save_to_db=False         # NEW: Don't auto-save to database
            )
        else:
            # Old FlightAnalyzer without database support
            results = flight_analyzer.analyze_flight(
                excel_path=excel_path,
                sheet_name=sheet_name,
                add_to_training=False  # Don't auto-add to training
            )
        
        if 'error' in results:
            return jsonify({'error': results['error']}), 500
        
        # Add analysis timestamp and status flags
        results['analysis_date'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        results['added_to_training'] = False  # NEW: User hasn't added yet
        results['saved_to_database'] = False  # NEW: User hasn't saved yet
        
        # Store in GLOBAL only (not session - too large for cookies!)
        # Session cookies have 4KB limit, visualization data is 100KB+
        latest_anomaly_report = results
        
        # Store only essential info in session (no visualization_data)
        session['analysis_metadata'] = {
            'flight_id': results.get('flight_id'),
            'analysis_date': results['analysis_date'],
            'added_to_training': False,
            'saved_to_database': False
        }
        
        # NEW: Store Excel path for later exceedances extraction
        session['excel_path'] = excel_path
        session['excel_filename'] = excel_filename
        
        session.modified = True
        
        print(f"✓ Analysis complete for Flight {results.get('flight_id', 'unknown')}")
        print(f"{'='*60}\n")
        
        # Return success - frontend will redirect to /anomaly_report
        return jsonify({'success': True})
        
    except Exception as e:
        print(f"❌ Error in flight anomaly analysis: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route("/anomaly_report", methods=["GET"])
def view_anomaly_report():
    """
    UPDATED: Display the anomaly report with new action buttons.
    Uses global variable (not session - too large for cookies).
    """
    global latest_anomaly_report
    
    # Use global variable for main data
    report_data = latest_anomaly_report
    
    if not report_data:
        return "No anomaly report available. Please run an analysis first.", 404
    
    # Get status flags from session if available
    session_metadata = session.get('analysis_metadata', {})
    
    # Update status flags from session (user actions)
    if session_metadata:
        report_data['added_to_training'] = session_metadata.get('added_to_training', False)
        report_data['saved_to_database'] = session_metadata.get('saved_to_database', False)
    
    # Get flight metadata from session
    flight_metadata = session.get('flight_metadata', None)
    
    # Prepare enhanced report data for template
    enhanced_report = {
        'flight_id': report_data.get('flight_id'),
        'analysis_date': report_data.get('analysis_date', datetime.now().strftime('%Y-%m-%d %H:%M:%S')),
        'total_data_points': report_data.get('total_data_points', 0),
        'anomaly_count': report_data.get('total_anomalies', 0),
        'anomaly_percentage': float(report_data.get('anomaly_percentage', 0.0)),  # Ensure it's a float
        'anomalies': report_data.get('anomalies', []),
        'anomalies_by_param_phase': report_data.get('anomalies_by_param_phase', {}),  # NEW
        'phases_summary': report_data.get('phases_summary', {}),
        'visualization_data': report_data.get('visualization_data', {}),
        'flight_metadata': flight_metadata,  # NEW
        'added_to_training': report_data.get('added_to_training', False),  # NEW
        'saved_to_database': report_data.get('saved_to_database', False),  # NEW
        'database_enabled': DATABASE_ENABLED,  # NEW - Show database button only if available
        'total_historical_flights': (
            flight_analyzer.historical_data['flight_id'].nunique()
            if hasattr(flight_analyzer, 'historical_data') and not flight_analyzer.historical_data.empty
            else 0
        )
    }
    
    # DEBUG: Print phases_summary structure to help diagnose issues
    print("\n🔍 DEBUG - phases_summary structure:")
    for phase, data in enhanced_report['phases_summary'].items():
        print(f"  Phase: {phase}")
        print(f"    Keys: {list(data.keys())}")
        print(f"    Values: {data}")
    
    return render_template("anomaly_report.html", report=enhanced_report)


@app.route("/", methods=["GET", "POST"])
def index():
    """
    UPDATED: Handle file upload and compliance checking.
    Now stores flight metadata in session and generates download URL.
    """
    if request.method == "POST":
        temp_request_dir = tempfile.mkdtemp()
        excel_file_path = None
        final_excel_output_path = None
        concatenated_audio_path = None
        cleaned_audio_path = None

        try:
            # Handle Excel file upload
            if 'excel_file' not in request.files:
                raise ValueError("No Excel file part in the request.")
            
            excel_file_upload = request.files['excel_file']
            if excel_file_upload.filename == '':
                raise ValueError("No selected Excel file.")
            
            excel_filename = secure_filename(excel_file_upload.filename)
            excel_file_path = os.path.join(temp_request_dir, excel_filename)
            excel_file_upload.save(excel_file_path)
            print(f"Excel file saved temporarily at: {excel_file_path}")

            # Handle Audio file(s) upload
            if 'audio_files[]' not in request.files:
                raise ValueError("No audio files part in the request.")
            
            uploaded_audio_files = request.files.getlist('audio_files[]')
            if not uploaded_audio_files or uploaded_audio_files[0].filename == '':
                raise ValueError("No audio files selected.")

            saved_audio_paths = []
            for file in uploaded_audio_files:
                if file:
                    audio_filename = secure_filename(file.filename)
                    audio_file_path = os.path.join(temp_request_dir, audio_filename)
                    file.save(audio_file_path)
                    saved_audio_paths.append(audio_file_path)

            if not saved_audio_paths:
                raise ValueError("No valid audio files uploaded.")

            # Get form data
            output_file_name = request.form.get("output_file_name", "concatenated_audio.wav")
            threshold = int(request.form.get("threshold", 50))
            sheet_name = request.form.get("sheet_name")

            if not sheet_name:
                raise ValueError("Sheet name is required.")
            if not output_file_name:
                raise ValueError("Output file name is required.")

            # NEW: Extract flight metadata from Excel file and filename
            print("\n📋 Extracting flight metadata...")
            flight_metadata = extract_flight_metadata_from_excel(excel_file_path, excel_filename)
            
            # Override with form data if provided (optional form fields)
            if request.form.get('flight_date'):
                flight_metadata['flight_date'] = request.form.get('flight_date')
            if request.form.get('pic'):
                flight_metadata['pic'] = request.form.get('pic').upper()
            if request.form.get('sic'):
                flight_metadata['sic'] = request.form.get('sic').upper()
            if request.form.get('fe'):
                flight_metadata['fe'] = request.form.get('fe').upper()
            if request.form.get('sortie'):
                flight_metadata['sortie'] = int(request.form.get('sortie'))
            if request.form.get('aircraft_id'):
                flight_metadata['aircraft_id'] = int(request.form.get('aircraft_id'))
            
            # Store in session
            session['flight_metadata'] = flight_metadata
            print(f"✓ Flight metadata: {flight_metadata}")

            # Audio Processing
            if len(saved_audio_paths) == 1:
                concatenated_audio_path = saved_audio_paths[0]
                print(f"Using single audio file: {os.path.basename(concatenated_audio_path)}")
            else:
                concatenated_audio_path = concatenate_audio_files(
                    saved_audio_paths, output_file_name, app.config["UPLOAD_FOLDER"]
                )
                if not concatenated_audio_path:
                    raise Exception("Audio concatenation failed.")
                print(f"Concatenated audio saved: {os.path.basename(concatenated_audio_path)}")

            cleaned_audio_path = preprocess_audio(concatenated_audio_path)
            if not cleaned_audio_path:
                raise Exception("Audio preprocessing failed.")

            # Transcribe audio
            print("Transcribing audio with Whisper...")
            transcript = transcribe_audio(cleaned_audio_path, output_file_name)
            print(f"Transcription complete: {len(transcript.split())} words")

            # Load Checklist and Check Compliance
            df, checklist, row_positions = load_checklist(excel_file_path, sheet_name)
            print(f"Checking compliance against {len(checklist)} checklist items...")
            results = check_compliance(transcript, checklist, threshold)
            
            # Map results to Excel row positions for database storage
            results_with_positions = []
            for i, (status, item, score, matched_text) in enumerate(results):
                excel_row = row_positions.get(i, i + 2)  # Fallback to i+2 if missing
                results_with_positions.append((status, item, score, matched_text, excel_row))
            
            print(f"  ✓ Mapped {len(results_with_positions)} results to Excel row positions")

            # Calculate compliance statistics
            passed_count = sum(1 for r in results if r[0] == "PASS")
            total_checks = len(results)
            compliance_percent = round((passed_count / total_checks) * 100, 1) if total_checks else 0
            not_complied_count = total_checks - passed_count

            # Update Excel with compliance results
            final_excel_output_path = update_excel(
                excel_file_path, results, sheet_name, not_complied_count, compliance_percent
            )

            # Save compliance report
            save_compliance_report(results, output_file_name)

            # FIXED: Re-extract metadata from the UPDATED Excel file
            # This ensures we get the correct compliance data that was just written
            print("\n📋 Re-extracting flight metadata from updated Excel file...")
            flight_metadata_updated = extract_flight_metadata_from_excel(
                final_excel_output_path, 
                os.path.basename(final_excel_output_path)
            )
            
            # Preserve any form overrides from the initial extraction
            for key in ['flight_date', 'pic', 'sic', 'fe', 'sortie', 'aircraft_id']:
                if key in flight_metadata and flight_metadata[key] != 'UNK':
                    flight_metadata_updated[key] = flight_metadata[key]
            
            # Update session with the corrected metadata
            flight_metadata = flight_metadata_updated
            session['flight_metadata'] = flight_metadata
            
            # Also store CVR results for later database save
            # Map sheet_name to checklist_type_id
            checklist_type_map = {
                'STARTING WITH AC-GPU CHECKLIST': 1,
                'STARTING WITH DC-GPU CHECKLIST': 2,
                'STARTING WITHOUT GPU CHECKLIST': 3
            }
            
            checklist_type_id = checklist_type_map.get(sheet_name, 1)
            
            session['cvr_results'] = {
                'results': [(r[0], r[1], r[2], r[3], r[4]) for r in results_with_positions],  # (status, item, score, matched_text, excel_row)
                'compliance_percent': compliance_percent,
                'not_complied_count': not_complied_count,
                'checklist_type_id': checklist_type_id,  # Store which checklist was used
                'sheet_name': sheet_name  # Store sheet name for reference
            }
            
            print(f"✓ Updated flight metadata with compliance data from Excel:")
            print(f"  - Compliance: {compliance_percent}%")
            print(f"  - Checks failed: {not_complied_count}")
            print(f"  - Checklist type: {sheet_name} (ID: {checklist_type_id})")

            # Return results with Excel filename for later anomaly analysis
            updated_excel_filename = os.path.basename(final_excel_output_path)
            
            # NEW: Generate download URL for the updated Excel file
            download_url = url_for('download_updated_excel', filename=updated_excel_filename, _external=True)

            print(f"\n✓ Compliance check complete:")
            print(f"  - Overall compliance: {compliance_percent}%")
            print(f"  - Checks failed: {not_complied_count}")
            print(f"  - Updated Excel: {updated_excel_filename}\n")

            return jsonify({
                "results": results,
                "compliance_percent": compliance_percent,
                "not_complied_count": not_complied_count,
                "excel_updated": True,
                "updated_excel_filename": updated_excel_filename,
                "download_excel_url": download_url,  # NEW
                "sheet_name": sheet_name
            })

        except Exception as e:
            print(f"❌ An error occurred: {e}")
            import traceback
            traceback.print_exc()
            return jsonify({"error": f"Error processing files: {e}"}), 500
        finally:
            # Clean up temporary directory
            if os.path.exists(temp_request_dir):
                shutil.rmtree(temp_request_dir)

    return render_template("index.html")


if __name__ == "__main__":
    print("\n" + "="*60)
    print("🚁 MI-17 Flight Analysis System Starting...")
    print("="*60)
    print(f"Database support: {'✓ Enabled' if DATABASE_ENABLED else '✗ Disabled'}")
    print(f"Upload folder: {UPLOAD_FOLDER}")
    print(f"Flight data folder: {FLIGHT_DATA_FOLDER}")
    print(f"Historical flights loaded: {flight_analyzer.historical_data['flight_id'].nunique() if hasattr(flight_analyzer, 'historical_data') and not flight_analyzer.historical_data.empty else 0}")
    print("="*60 + "\n")
    
    app.run(debug=True, host='0.0.0.0', port=5000)