import os
import pandas as pd
from flask import Flask, render_template, request, redirect, url_for, jsonify
from werkzeug.utils import secure_filename
from rapidfuzz import fuzz
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.styles import Alignment
import subprocess
from statistics import mean
import re
from faster_whisper import WhisperModel
import tempfile
import shutil
from datetime import datetime

# Flask setup
app = Flask(__name__)
UPLOAD_FOLDER = "uploads"
app.config["UPLOAD_FOLDER"] = UPLOAD_FOLDER
EXCEL_FILE = "workbench/WORKBENCH_CARE.xlsm"
CHECKED_COLUMN = "B"
TRANSCRIPT_FOLDER = "transcripts"
COMPLIANCE_FOLDER = "compliance" # New folder for compliance reports

# Ensure transcript folder exists
os.makedirs(TRANSCRIPT_FOLDER, exist_ok=True)
# Ensure upload folder exists
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
# Ensure workbench folder exists for the excel file
os.makedirs("workbench", exist_ok=True)
# Ensure compliance folder exists
os.makedirs(COMPLIANCE_FOLDER, exist_ok=True)


# Initialize Whisper model globally (to avoid reloading for every request)
model = WhisperModel("medium", device="cuda", compute_type="float16")

def preprocess_audio(input_path, output_folder="uploads", filename_prefix="cleaned_"):
    """
    Converts, trims silence, normalizes volume, and prepares audio for transcription.
    Returns path to the cleaned WAV file.
    """
    os.makedirs(output_folder, exist_ok=True)

    base = os.path.basename(input_path)
    name, _ = os.path.splitext(base)
    output_path = os.path.join(output_folder, f"{filename_prefix}{name}.wav")

    command = [
        "ffmpeg", "-y",
        "-i", input_path,
        "-ac", "1",
        "-ar", "16000",
        "-af", "silenceremove=1:0:-50dB,loudnorm",
        "-c:a", "pcm_s16le",
        output_path
    ]

    try:
        # Capture stderr to see FFmpeg's error messages
        result = subprocess.run(command, capture_output=True, text=True, check=True)
        print(f"FFmpeg stdout (preprocessing): {result.stdout}")
        return output_path
    except subprocess.CalledProcessError as e:
        print(f"FFmpeg failed during preprocessing: {e}")
        print(f"FFmpeg stderr (preprocessing): {e.stderr}")
        return None

def concatenate_audio_files(input_paths, output_filename, upload_folder):
    """
    Concatenates multiple audio files into a single WAV file using ffmpeg.
    Returns the path to the concatenated file.
    """
    concat_list_path = os.path.join(tempfile.gettempdir(), "files.txt")

    with open(concat_list_path, "w") as f:
        for path in input_paths:
            f.write(f"file '{path}'\n")

    # Ensure output_filename has a .wav extension
    if not output_filename.lower().endswith(".wav"):
        output_filename += ".wav"

    output_path = os.path.join(upload_folder, secure_filename(output_filename))

    command = [
        "ffmpeg", "-y",
        "-f", "concat",
        "-safe", "0", # Required for external file paths in concat list
        "-i", concat_list_path,
        "-c", "copy", # Use stream copy for maximum speed if codecs are compatible
        output_path
    ]

    try:
        result = subprocess.run(command, capture_output=True, text=True, check=True)
        print(f"FFmpeg stdout (concatenation): {result.stdout}")
        return output_path
    except subprocess.CalledProcessError as e:
        print(f"FFmpeg failed during concatenation: {e}")
        print(f"FFmpeg stderr (concatenation): {e.stderr}")
        return None
    finally:
        # Clean up the temporary concat list file
        if os.path.exists(concat_list_path):
            os.remove(concat_list_path)

def load_checklist(sheet_name):
    """
    Loads the checklist items from the specified Excel sheet.
    Assumes checklist items are in the first column.
    """
    df = pd.read_excel(EXCEL_FILE, sheet_name=sheet_name, engine="openpyxl")
    return df, df.iloc[:, 0].dropna().tolist()

def clean_text(text):
    """
    Cleans text by converting to lowercase, removing non-alphanumeric characters,
    and removing common filler words.
    """
    text = text.lower()
    text = re.sub(r"[^a-zA-Z0-9\s]", "", text)
    text = re.sub(r"\b(?:roger|copy|standby|okay|affirmative|negative|check)\b", "", text)
    return text.strip()

def check_compliance(transcript, checklist, threshold=50):
    """
    Checks compliance of the transcript against the checklist using fuzzy matching.
    This version uses a sliding window across the entire transcript (word-by-word)
    to find the best possible match for each checklist item.
    """
    transcript_lower = transcript.lower()
    transcript_words_raw = transcript_lower.split()

    results = []
    MAX_CHUNK_WORDS = 20

    for step in checklist:
        step_clean = clean_text(step)
        best_score = 0
        best_chunk_raw = ""

        for i in range(len(transcript_words_raw)):
            for j in range(i + 1, min(i + MAX_CHUNK_WORDS + 1, len(transcript_words_raw) + 1)):
                current_chunk_words_raw = transcript_words_raw[i:j]
                current_chunk_raw = ' '.join(current_chunk_words_raw)

                current_chunk_clean = clean_text(current_chunk_raw)

                if not current_chunk_clean:
                    continue

                pr = fuzz.partial_ratio(step_clean, current_chunk_clean)
                tsr = fuzz.token_set_ratio(step_clean, current_chunk_clean)
                ratio = fuzz.ratio(step_clean, current_chunk_clean)

                score = max(pr, tsr, ratio) * 0.6 + mean([pr, tsr, ratio]) * 0.4

                if score > best_score:
                    best_score = score
                    best_chunk_raw = current_chunk_raw

        if best_score == 100.0 and step_clean not in clean_text(transcript):
            best_score = 99.0

        # Print to console (for backend debugging/logging)
        print(f"\n✅ Checklist Item: {step}")
        print(f"   🔍 Matched: \"{best_chunk_raw}\"")
        print(f"   🎯 Score: {best_score:.1f}%")

        results.append(("PASS" if best_score >= threshold else "FAIL", step, best_score, best_chunk_raw)) # Added best_chunk_raw

    return results

def update_excel(results, sheet_name, not_complied_count, compliance_percent):
    """
    Updates the Excel file with compliance results.
    Marks cells in CHECKED_COLUMN with a checkmark (✔) for PASS and cross (✘) for FAIL,
    with corresponding green/red colors. Also adds non-complied count and complied percentage.
    """
    wb = load_workbook(EXCEL_FILE, keep_vba=True)
    ws = wb[sheet_name]

    # Update checklist results
    row = 2
    for result in results:
        status_icon = "✔" if result[0] == "PASS" else "✘"
        cell = ws[f"{CHECKED_COLUMN}{row}"]
        cell.value = status_icon
        cell.font = Font(color="008000" if result[0] == "PASS" else "FF0000")
        row += 1

    # Apply professional formatting for "Checks Not Complied"
    ws['D1'].value = "Checks Not Complied:"
    ws['D1'].font = Font(bold=True)
    ws['E1'].value = not_complied_count    
    ws['E1'].font = Font(bold=True, color="FF0000") # Red color for not complied count
    ws['E1'].fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid") # Light orange background
    ws['E1'].alignment = Alignment(horizontal='center', vertical='center')
    # Apply professional formatting for "Complied Percentage"
    ws['D2'].value = "Complied Percentage:"
    ws['D2'].font = Font(bold=True)
    ws['E2'].value = f"{compliance_percent:.1f}%"
    ws['E2'].font = Font(bold=True, color="008000") # Green color for complied percentage
    ws['E2'].fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid") # Light green background
    ws['E2'].alignment = Alignment(horizontal='center', vertical='center')

    wb.save(EXCEL_FILE)

def transcribe_audio(audio_path):
    """
    Transcribes audio using the FasterWhisper model and saves the transcript to a file.
    """
    segments, info = model.transcribe(audio_path, language="en")

    transcript_text = " ".join([segment.text for segment in segments])

    # Save transcript
    base_filename = os.path.splitext(os.path.basename(audio_path))[0]
    transcript_filename = f"{base_filename}.txt"
    transcript_path = os.path.join(TRANSCRIPT_FOLDER, transcript_filename)

    with open(transcript_path, "w", encoding="utf-8") as f:
        f.write(transcript_text)

    return transcript_text

def save_compliance_report(results, output_file_name):
    """
    Saves the compliance results to a text file in the compliance folder.
    """
    # Sanitize output_file_name for use in filename, remove extension if present
    base_name = os.path.splitext(secure_filename(output_file_name))[0]

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    report_filename = f"{base_name}_compliance_report_{timestamp}.txt"
    report_path = os.path.join(COMPLIANCE_FOLDER, report_filename)

    with open(report_path, "w", encoding="utf-8") as f:
        f.write(f"Compliance Report for: {output_file_name}\n")
        f.write(f"Generated On: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        f.write("-" * 50 + "\n\n")

        for status, checklist_item, score, matched_text in results:
            f.write(f"Status: {status}\n")
            f.write(f"Checklist Item: {checklist_item}\n")
            f.write(f"Matched Text: \"{matched_text}\"\n") # Include matched text
            f.write(f"Score: {score:.1f}%\n")
            f.write("-" * 20 + "\n")
    print(f"Compliance report saved to: {report_path}")


# Routes
@app.route("/", methods=["GET", "POST"])
def index():
    """
    Handles file upload, audio preprocessing, transcription, compliance checking,
    and returns the report as JSON.
    """
    if request.method == "POST":
        if 'files[]' not in request.files:
            return jsonify({"error": "No files part"}), 400

        uploaded_files = request.files.getlist('files[]')
        output_file_name = request.form.get("output_file_name", "concatenated_audio.wav")
        threshold = int(request.form.get("threshold", 50))
        sheet_name = request.form.get("sheet_name")

        if not uploaded_files or uploaded_files[0].filename == '':
            return jsonify({"error": "No selected file"}), 400

        if not sheet_name:
            return jsonify({"error": "Sheet name is required"}), 400

        temp_upload_dir = tempfile.mkdtemp()
        saved_file_paths = []
        concatenated_audio_path = None
        cleaned_path = None

        try:
            for file in uploaded_files:
                if file:
                    filename = secure_filename(file.filename)
                    file_path = os.path.join(temp_upload_dir, filename)
                    file.save(file_path)
                    saved_file_paths.append(file_path)

            if not saved_file_paths:
                return jsonify({"error": "No valid files uploaded"}), 400

            # Optimization: Skip concatenation if only one file is uploaded
            if len(saved_file_paths) == 1:
                # If only one file, use it directly as the source for preprocessing
                concatenated_audio_path = saved_file_paths[0]
                print(f"Skipping concatenation: Directly processing single file: {os.path.basename(concatenated_audio_path)}")
            else:
                # If multiple files, proceed with concatenation
                concatenated_audio_path = concatenate_audio_files(saved_file_paths, output_file_name, app.config["UPLOAD_FOLDER"])
                if not concatenated_audio_path:
                    return jsonify({"error": "Audio concatenation failed. Check FFmpeg installation and file formats."}), 500
                print(f"Concatenated audio saved to: {os.path.basename(concatenated_audio_path)}")


            # Preprocess the concatenated (or single) audio file
            cleaned_path = preprocess_audio(concatenated_audio_path)
            if not cleaned_path:
                return jsonify({"error": "Audio preprocessing failed"}), 500

            # Load checklist and transcribe cleaned audio
            df, checklist = load_checklist(sheet_name)
            transcript = transcribe_audio(cleaned_path)

            # Check compliance
            results = check_compliance(transcript, checklist, threshold)

            # Calculate compliance statistics for frontend and Excel
            passed_count = sum(1 for r in results if r[0] == "PASS")
            total_checks = len(results)
            compliance_percent = round((passed_count / total_checks) * 100, 1) if total_checks else 0

            not_complied_count = total_checks - passed_count

            # Update Excel
            try:
                update_excel(results, sheet_name, not_complied_count, compliance_percent)
            except Exception as e:
                return jsonify({"error": f"Error updating Excel file: {e}. Please ensure the Excel file is not open and accessible."}), 500

            # Save compliance report to file
            save_compliance_report(results, output_file_name)

            # Return results as JSON for frontend
            return jsonify({
                "results": results,
                "compliance_percent": compliance_percent,
                "not_complied_count": not_complied_count
            })

        except Exception as e:
            print(f"An unexpected error occurred: {e}")
            return jsonify({"error": f"An unexpected error occurred: {e}"}), 500
        finally:
            # Clean up the temporary directory after processing
            if os.path.exists(temp_upload_dir):
                shutil.rmtree(temp_upload_dir)

            # Remove the cleaned audio file (it's temporary for transcription)
            if cleaned_path and os.path.exists(cleaned_path):
                os.remove(cleaned_path)

            # If concatenation occurred, you might want to keep the concatenated file
            # or remove it based on your desired behavior. For now, it remains in 'uploads'.
            # If you want to remove it:
            # if concatenated_audio_path and os.path.exists(concatenated_audio_path) and len(saved_file_paths) > 1:
            #     os.remove(concatenated_audio_path)


    # For GET requests, render the React app's entry point
    return render_template("index.html")

# Run Flask app
if __name__ == "__main__":
    app.run(debug=True, host='0.0.0.0', port=5000)